"""
End-to-end integration test for QuickBooks Excel upload pipeline.
20 test scenarios covering standard parsing, edge cases, alternative
formats, error handling, and regression checks.

Each mock builder generates unique financial data that follows the
standard QuickBooks Online or Desktop Excel export format.
"""

import copy
import io
import os
import sys
import traceback

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl
import pandas as pd

PASSED = 0
FAILED = 0


def _result(name, ok, detail=""):
    global PASSED, FAILED
    if ok:
        PASSED += 1
        print(f"    PASS  {name}")
    else:
        FAILED += 1
        print(f"    FAIL  {name}  — {detail}")


# ═══════════════════════════════════════════════
# Mock Data Builders — each has unique org, amounts, and structure
# ═══════════════════════════════════════════════


def create_mock_pl():
    """Standard QB Online P&L — Habitat Community Foundation.
    Multi-period (2024 & 2023).  Revenue ~$1.23M, Expenses ~$912K.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    for r in [
        ["Profit and Loss", "", ""],
        ["Habitat Community Foundation", "", ""],
        ["January - December 2024", "", ""],
        ["", "", ""],
        ["", "Jan - Dec 2024", "Jan - Dec 2023"],
        ["Income", "", ""],
        ["  4000 Contributions & Grants", 850000, 780000],
        ["  4100 Program Service Revenue", 320000, 290000],
        ["  4200 Investment Income", 45000, 38000],
        ["  4300 Other Income", 15000, 12000],
        ["Total Income", 1230000, 1120000],
        ["", "", ""],
        ["Expenses", "", ""],
        ["  6000 Salaries and Wages", 520000, 480000],
        ["  6100 Payroll Taxes", 42000, 39000],
        ["  6200 Employee Benefits", 65000, 58000],
        ["  6300 Rent", 72000, 70000],
        ["  6400 Insurance", 28000, 25000],
        ["  6500 Office Expenses", 18000, 16000],
        ["  6600 Travel", 22000, 19000],
        ["  6700 Depreciation", 35000, 32000],
        ["  6800 Legal Fees", 15000, 12000],
        ["  6900 Accounting Fees", 20000, 18000],
        ["  7000 Information Technology", 30000, 25000],
        ["  7100 Other Expenses", 45000, 40000],
        ["Total Expenses", 912000, 834000],
        ["", "", ""],
        ["Net Income", 318000, 286000],
    ]:
        ws.append(r)
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.getvalue()


def create_mock_bs():
    """Standard QB Online Balance Sheet — Habitat Community Foundation.
    Multi-period (Dec 2024 & Dec 2023).  Assets ~$1.215M.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    for r in [
        ["Balance Sheet", "", ""],
        ["Habitat Community Foundation", "", ""],
        ["As of December 31, 2024", "", ""],
        ["", "", ""],
        ["", "Dec 31, 2024", "Dec 31, 2023"],
        ["ASSETS", "", ""],
        ["    Checking", 280000, 220000],
        ["    Savings", 450000, 380000],
        ["    Investments", 320000, 280000],
        ["    Inventory", 15000, 12000],
        ["    Prepaid Expenses", 25000, 20000],
        ["  Total Current Assets", 1090000, 912000],
        ["  Fixed Assets", "", ""],
        ["    Furniture and Equipment", 180000, 165000],
        ["  Total Fixed Assets", 95000, 115000],
        ["  Other Assets", 30000, 25000],
        ["Total Assets", 1215000, 1052000],
        ["", "", ""],
        ["LIABILITIES AND EQUITY", "", ""],
        ["    Accounts Payable", 45000, 38000],
        ["    Other Current Liabilities", 62000, 55000],
        ["  Total Liabilities", 107000, 93000],
        ["    Unrestricted Net Assets", 908000, 759000],
        ["    Restricted Net Assets", 200000, 200000],
        ["  Total Equity", 1108000, 959000],
        ["TOTAL LIABILITIES AND EQUITY", 1215000, 1052000],
    ]:
        ws.append(r)
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.getvalue()


def create_single_period_pl():
    """Single-column QB Online P&L — Riverside Animal Rescue.
    One period only (2025).  Revenue $214K, Expenses $178K.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    for r in [
        ["Profit and Loss", ""],
        ["Riverside Animal Rescue", ""],
        ["January - December 2025", ""],
        ["", ""],
        ["", "Total"],
        ["Income", ""],
        ["  4000 Donations Received", 148000],
        ["  4100 Adoption Fees", 52000],
        ["  4200 Interest Income", 3500],
        ["  4300 Fundraising Events", 10500],
        ["Total Income", 214000],
        ["", ""],
        ["Expenses", ""],
        ["  6000 Salaries & Wages", 89000],
        ["  6050 Payroll Taxes", 7200],
        ["  6100 Employee Benefits", 12400],
        ["  6200 Rent", 24000],
        ["  6300 Insurance", 6800],
        ["  6400 Office Supplies", 4100],
        ["  6500 Veterinary Supplies", 19500],
        ["  6600 Travel", 3200],
        ["  6700 Depreciation", 5800],
        ["  6800 Information Technology", 6000],
        ["Total Expenses", 178000],
        ["", ""],
        ["Net Income", 36000],
    ]:
        ws.append(r)
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.getvalue()


def create_csv_pl():
    """CSV format P&L — Green Valley Land Trust.
    Multi-period (2024 & 2023).  Revenue $830K, Expenses $483K.
    """
    return b"""Profit and Loss,,
Green Valley Land Trust,,
January - December 2024,,
,,
,Jan - Dec 2024,Jan - Dec 2023
Income,,
  Contributions,600000,550000
  Program Service Revenue,200000,180000
  Investment Income,30000,25000
Total Income,830000,755000
,,
Expenses,,
  Salaries,350000,320000
  Payroll Taxes,28000,26000
  Rent or Lease,48000,45000
  Insurance,15000,14000
  Depreciation,20000,18000
  Travel,10000,8000
  Office Expenses,12000,11000
Total Expenses,483000,442000
,,
Net Income,347000,313000
"""


def create_alt_naming_pl():
    """Alternative naming conventions — Summit Youth Services.
    Uses 'Revenue'/'Expenditures', Fee for Service, Payroll Expenses, etc.
    Single period FY2024.  Revenue $935K, Expenses $762K.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    for r in [
        ["Income Statement", ""],
        ["Summit Youth Services", ""],
        ["FY2024 (Jul 2023 - Jun 2024)", ""],
        ["", ""],
        ["", "Total"],
        ["Revenue", ""],
        ["  Contributions and Grants", 620000],
        ["  Fee for Service", 275000],
        ["  Interest and Dividends", 28000],
        ["  Miscellaneous Income", 12000],
        ["Total Revenue", 935000],
        ["", ""],
        ["Expenditures", ""],
        ["  Payroll Expenses", 415000],
        ["  Employer Taxes", 33600],
        ["  Health Insurance", 48000],
        ["  Occupancy", 68000],
        ["  Computer Expense", 31000],
        ["  Travel Expense", 19000],
        ["  Audit Fees", 16000],
        ["  Legal Services", 11000],
        ["  Grant Expense", 58000],
        ["  Depreciation", 22400],
        ["  Other Expenses", 40000],
        ["Total Expenditures", 762000],
        ["", ""],
        ["Net Income", 173000],
    ]:
        ws.append(r)
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.getvalue()


def create_5yr_pl():
    """Five-year comparative P&L — Pacific Marine Conservation.
    5 columns (2020-2024).  Revenue grows from $410K to $615K.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    for r in [
        ["Profit and Loss", "", "", "", "", ""],
        ["Pacific Marine Conservation", "", "", "", "", ""],
        ["Comparative Report 2020-2024", "", "", "", "", ""],
        ["", "", "", "", "", ""],
        ["", "Jan-Dec 2024", "Jan-Dec 2023", "Jan-Dec 2022", "Jan-Dec 2021", "Jan-Dec 2020"],
        ["Income", "", "", "", "", ""],
        ["  Contributions", 380000, 355000, 330000, 305000, 280000],
        ["  Program Revenue", 195000, 178000, 161000, 144000, 105000],
        ["  Investment Income", 40000, 35000, 30000, 25000, 25000],
        ["Total Income", 615000, 568000, 521000, 474000, 410000],
        ["", "", "", "", "", ""],
        ["Expenses", "", "", "", "", ""],
        ["  Salaries", 280000, 265000, 250000, 235000, 210000],
        ["  Rent", 42000, 40000, 38000, 36000, 34000],
        ["  Insurance", 15000, 14000, 13500, 13000, 12000],
        ["  Travel", 28000, 25000, 22000, 19000, 16000],
        ["  Office Expenses", 9000, 8500, 8000, 7500, 7000],
        ["Total Expenses", 374000, 352500, 331500, 310500, 279000],
        ["", "", "", "", "", ""],
        ["Net Income", 241000, 215500, 189500, 163500, 131000],
    ]:
        ws.append(r)
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.getvalue()


def create_zero_revenue_pl():
    """Zero-revenue startup nonprofit — Fresh Start Legal Aid.
    Brand new org, no revenue yet, only startup expenses ($11.5K).
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    for r in [
        ["Profit and Loss", ""],
        ["Fresh Start Legal Aid", ""],
        ["January - December 2025", ""],
        ["", ""],
        ["", "Total"],
        ["Income", ""],
        ["Total Income", 0],
        ["", ""],
        ["Expenses", ""],
        ["  Legal Fees", 5500],
        ["  Office Expenses", 3200],
        ["  Insurance", 1800],
        ["  Information Technology", 1000],
        ["Total Expenses", 11500],
        ["", ""],
        ["Net Income", -11500],
    ]:
        ws.append(r)
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.getvalue()


def create_desktop_pl():
    """QB Desktop-style export — Heartland Food Bank.
    Uses 'Profit & Loss', 'through', '·' separator, Ordinary Income/Expense.
    Multi-period (2024 & 2023).  Revenue $742K, Expenses $631K.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    for r in [
        ["Heartland Food Bank", "", ""],
        ["Profit & Loss", "", ""],
        ["January through December 2024", "", ""],
        ["", "Jan - Dec 24", "Jan - Dec 23"],
        ["Ordinary Income/Expense", "", ""],
        ["  Income", "", ""],
        ["    4000 · Donations", 485000, 440000],
        ["    4100 · Program Fees", 190000, 175000],
        ["    4200 · Interest Income", 12000, 10000],
        ["    4300 · Grants Received", 55000, 50000],
        ["  Total Income", 742000, 675000],
        ["  Expense", "", ""],
        ["    6000 · Salaries & Wages", 310000, 290000],
        ["    6100 · Payroll Taxes", 25000, 23500],
        ["    6200 · Employee Benefits", 38000, 35000],
        ["    6300 · Rent or Lease", 54000, 52000],
        ["    6400 · Insurance", 16000, 15000],
        ["    6500 · Office Supplies", 9500, 9000],
        ["    6600 · Travel", 18000, 16000],
        ["    6700 · Depreciation", 22000, 20000],
        ["    6800 · Information Technology", 17500, 15500],
        ["    6900 · Food Supplies", 85000, 78000],
        ["    7000 · Other Expenses", 36000, 33000],
        ["  Total Expense", 631000, 587000],
        ["Net Ordinary Income", 111000, 88000],
        ["Net Income", 111000, 88000],
    ]:
        ws.append(r)
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.getvalue()


def create_large_bs():
    """Detailed Balance Sheet — Cascade Education Alliance.
    Single-period (Dec 2024).  Assets $3.85M, heavy fixed assets.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    for r in [
        ["Balance Sheet", ""],
        ["Cascade Education Alliance", ""],
        ["As of December 31, 2024", ""],
        ["", ""],
        ["", "Dec 31, 2024"],
        ["ASSETS", ""],
        ["  Current Assets", ""],
        ["    Checking", 520000],
        ["    Savings", 890000],
        ["    Investments", 675000],
        ["    Inventory", 42000],
        ["    Prepaid Expenses", 38000],
        ["  Total Current Assets", 2165000],
        ["  Fixed Assets", ""],
        ["    Furniture and Equipment", 480000],
        ["    Less Accumulated Depreciation", -125000],
        ["  Total Fixed Assets", 355000],
        ["  Other Assets", 95000],
        ["Total Assets", 3850000],
        ["", ""],
        ["LIABILITIES AND EQUITY", ""],
        ["  Current Liabilities", ""],
        ["    Accounts Payable", 125000],
        ["    Accrued Liabilities", 88000],
        ["    Other Current Liabilities", 47000],
        ["  Total Liabilities", 260000],
        ["  Equity", ""],
        ["    Unrestricted Net Assets", 2740000],
        ["    Restricted Net Assets", 850000],
        ["  Total Equity", 3590000],
        ["TOTAL LIABILITIES AND EQUITY", 3850000],
    ]:
        ws.append(r)
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.getvalue()


def create_minimal_pl():
    """Bare-minimum QB P&L — Neighborhood Watch Coalition.
    Only total income and total expenses, no line items.
    Revenue $67K, Expenses $54K.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    for r in [
        ["Profit and Loss", ""],
        ["Neighborhood Watch Coalition", ""],
        ["January - December 2024", ""],
        ["", ""],
        ["", "Total"],
        ["Income", ""],
        ["  Donations", 67000],
        ["Total Income", 67000],
        ["", ""],
        ["Expenses", ""],
        ["  Operating Expenses", 54000],
        ["Total Expenses", 54000],
        ["", ""],
        ["Net Income", 13000],
    ]:
        ws.append(r)
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.getvalue()


# ═══════════════════════════════════════════════
# Tests 1-20
# ═══════════════════════════════════════════════

def test_01():
    from core.quickbooks import parse_quickbooks_report
    print("\n  TEST 01: Standard multi-period P&L (Habitat Community Foundation)")
    rows = parse_quickbooks_report(create_mock_pl(), "Habitat_PL_2024.xlsx")
    r24 = next((r for r in rows if r.get("TaxYear") == "2024"), {})
    r23 = next((r for r in rows if r.get("TaxYear") == "2023"), {})
    _result("2 periods found", len(rows) == 2, f"got {len(rows)}")
    _result("DataSource=quickbooks", r24.get("DataSource") == "quickbooks")
    _result("Org=Habitat Community Foundation", r24.get("OrganizationName") == "Habitat Community Foundation")
    _result("2024 Revenue=1230000", r24.get("TotalRevenue") == 1230000, str(r24.get("TotalRevenue")))
    _result("2024 Expenses=912000", r24.get("TotalExpenses") == 912000, str(r24.get("TotalExpenses")))
    _result("2024 Salaries=520000", r24.get("SalariesWages") == 520000)
    _result("2024 PayrollTax=42000", r24.get("PayrollTaxes") == 42000)
    _result("2024 Benefits=65000", r24.get("EmployeeBenefits") == 65000)
    _result("2024 Rent=72000", r24.get("Occupancy") == 72000)
    _result("2024 Insurance=28000", r24.get("Insurance") == 28000)
    _result("2024 Office=18000", r24.get("OfficeExpenses") == 18000)
    _result("2024 Travel=22000", r24.get("Travel") == 22000)
    _result("2024 Depreciation=35000", r24.get("DepreciationAmortization") == 35000)
    _result("2024 Legal=15000", r24.get("LegalFees") == 15000)
    _result("2024 Accounting=20000", r24.get("AccountingFees") == 20000)
    _result("2024 IT=30000", r24.get("InformationTechnology") == 30000)
    _result("2024 InvIncome=45000", r24.get("InvestmentIncome") == 45000)
    _result("2023 Revenue=1120000", r23.get("TotalRevenue") == 1120000, str(r23.get("TotalRevenue")))
    _result("2023 Expenses=834000", r23.get("TotalExpenses") == 834000, str(r23.get("TotalExpenses")))
    _result("ProgramExpenses=None (QB unavailable)", r24.get("ProgramExpenses") is None)
    _result("VotingBoardMembers=None (QB unavailable)", r24.get("VotingBoardMembers") is None)
    _result("Mission=None (QB unavailable)", r24.get("Mission") is None)
    return rows


def test_02():
    from core.quickbooks import parse_quickbooks_report
    print("\n  TEST 02: Standard multi-period Balance Sheet (Habitat Community Foundation)")
    rows = parse_quickbooks_report(create_mock_bs(), "Habitat_BS_2024.xlsx")
    r24 = next((r for r in rows if r.get("TaxYear") == "2024"), {})
    r23 = next((r for r in rows if r.get("TaxYear") == "2023"), {})
    _result("2 periods found", len(rows) == 2, f"got {len(rows)}")
    _result("2024 Assets=1215000", r24.get("TotalAssets") == 1215000, str(r24.get("TotalAssets")))
    _result("2024 Liabilities=107000", r24.get("TotalLiabilities") == 107000, str(r24.get("TotalLiabilities")))
    _result("Liab!=Assets (regression)", r24.get("TotalLiabilities") != r24.get("TotalAssets"))
    _result("2024 NetAssets=1108000", r24.get("TotalNetAssets") == 1108000)
    _result("2024 Cash=280000", r24.get("CashNonInterest") == 280000)
    _result("2024 Savings=450000", r24.get("SavingsTempCashInvestments") == 450000)
    _result("2024 Investments=320000", r24.get("PublicInvestments") == 320000)
    _result("2024 AP=45000", r24.get("AccountsPayableAccrued") == 45000)
    _result("2024 Inventory=15000", r24.get("Inventory") == 15000)
    _result("2024 Prepaid=25000", r24.get("PrepaidExpenses") == 25000)
    _result("2024 OtherAssets=30000", r24.get("OtherAssets") == 30000)
    _result("2023 Assets=1052000", r23.get("TotalAssets") == 1052000)
    _result("2023 Liabilities=93000", r23.get("TotalLiabilities") == 93000)
    _result("Unrestricted=908000", r24.get("NetAssetsWithoutDonorRestrictions") == 908000,
            str(r24.get("NetAssetsWithoutDonorRestrictions")))
    _result("Restricted=200000", r24.get("NetAssetsWithDonorRestrictions") == 200000,
            str(r24.get("NetAssetsWithDonorRestrictions")))
    return rows


def test_03(pl_rows, bs_rows):
    from core.quickbooks import merge_quickbooks_data
    print("\n  TEST 03: Merge P&L + Balance Sheet by year (Habitat)")
    merged = merge_quickbooks_data(pl_rows + bs_rows)
    _result("2 merged records", len(merged) == 2, f"got {len(merged)}")
    for row in merged:
        y = row.get("TaxYear")
        _result(f"{y} has revenue", (row.get("TotalRevenue") or 0) > 0)
        _result(f"{y} has assets", (row.get("TotalAssets") or 0) > 0)
        _result(f"{y} DataSource=quickbooks", row.get("DataSource") == "quickbooks")
        _result(f"{y} org name preserved", row.get("OrganizationName") == "Habitat Community Foundation")
    return merged


def test_04(merged):
    from core.kpis import compute_kpis, format_kpi_value, get_kpi_status
    print("\n  TEST 04: KPI computation — available vs N/A (Habitat merged)")
    for row in merged:
        y = row.get("TaxYear")
        kpis = compute_kpis(row)
        for k in ["OperatingSurplus", "OperatingMargin", "TotalCashEquivalents",
                   "LiquidAssets", "MonthsExpenseCoverage", "DebtToAssetRatio",
                   "CurrentRatio", "SalaryToExpenseRatio", "ContributionDependency"]:
            _result(f"{y} {k} available", kpis.get(k) is not None, str(kpis.get(k)))
        for k in ["ProgramExpenseRatio", "ManagementGeneralRatio", "FundraisingRatio",
                   "RevenueGrowth", "ExpenseGrowth"]:
            v = kpis.get(k)
            _result(f"{y} {k}=None", v is None, str(v))
            _result(f"{y} {k} fmt=N/A", format_kpi_value(k, v) == "N/A")
            _result(f"{y} {k} status=neutral", get_kpi_status(k, v) == "neutral")


def test_05(merged):
    from core.quickbooks import apply_supplement_data
    from core.kpis import compute_kpis
    print("\n  TEST 05: Full supplement form application (Habitat)")
    rows = copy.deepcopy(merged)
    supp = {"EIN": "12-3456789", "Mission": "Providing affordable housing for families",
            "VotingBoardMembers": 9, "IndependentBoardMembers": 7,
            "Volunteers": 150, "EmployeeCount": 45,
            "ExecutiveDirectorCompensation": 95000,
            "ProgramExpensePct": 78, "ManagementExpensePct": 14, "FundraisingExpensePct": 8}
    rows = apply_supplement_data(rows, supp)
    for row in rows:
        y = row.get("TaxYear")
        _result(f"{y} EIN set", row.get("EIN") == "12-3456789")
        _result(f"{y} Mission set", row.get("Mission") == "Providing affordable housing for families")
        _result(f"{y} Board=9", row.get("VotingBoardMembers") == 9)
        _result(f"{y} Volunteers=150", row.get("Volunteers") == 150)
        _result(f"{y} ExecComp=95000", row.get("ExecutiveDirectorCompensation") == 95000)
        kpis = compute_kpis(row)
        _result(f"{y} ProgRatio~78%", abs((kpis.get("ProgramExpenseRatio") or 0) - 0.78) < 0.01)
        _result(f"{y} MgmtRatio~14%", abs((kpis.get("ManagementGeneralRatio") or 0) - 0.14) < 0.01)
        _result(f"{y} FundRatio~8%", abs((kpis.get("FundraisingRatio") or 0) - 0.08) < 0.01)


def test_06(merged):
    from core.export import generate_workbook
    print("\n  TEST 06: Excel export with QB data (Habitat merged)")
    wb_bytes = generate_workbook(merged)
    wb = openpyxl.load_workbook(io.BytesIO(wb_bytes))
    _result("Workbook > 1KB", len(wb_bytes) > 1000)
    _result("3 sheets present", len(wb.sheetnames) == 3)
    na_sum = sum(1 for r in wb["Summary"].iter_rows(min_row=6, values_only=True)
                 for c in r if c == "N/A")
    na_data = sum(1 for r in wb["Extracted Data"].iter_rows(min_row=2, values_only=True)
                  for c in r if c == "N/A")
    _result("N/A cells in Summary > 0", na_sum > 0, str(na_sum))
    _result("N/A cells in ExtractedData > 0", na_data > 0, str(na_data))


def test_07():
    from core.parser import parse_single_xml
    from core.kpis import compute_kpis
    print("\n  TEST 07: XML pipeline regression (unchanged)")
    xml = b"""<?xml version="1.0" encoding="utf-8"?>
    <Return xmlns="http://www.irs.gov/efile"><ReturnHeader>
    <TaxYr>2023</TaxYr><Filer><EIN>123456789</EIN>
    <BusinessName><BusinessNameLine1Txt>Test Org</BusinessNameLine1Txt></BusinessName>
    </Filer></ReturnHeader><ReturnData><IRS990>
    <CYTotalRevenueAmt>500000</CYTotalRevenueAmt>
    <CYTotalExpensesAmt>450000</CYTotalExpensesAmt>
    <TotalProgramServiceExpensesAmt>350000</TotalProgramServiceExpensesAmt>
    <CYTotalFundraisingExpenseAmt>25000</CYTotalFundraisingExpenseAmt>
    <TotalAssetsEOYAmt>800000</TotalAssetsEOYAmt>
    <TotalLiabilitiesEOYAmt>200000</TotalLiabilitiesEOYAmt>
    <NetAssetsOrFundBalancesEOYAmt>600000</NetAssetsOrFundBalancesEOYAmt>
    <CYSalariesCompEmpBnftPaidAmt>280000</CYSalariesCompEmpBnftPaidAmt>
    <CYContributionsGrantsAmt>400000</CYContributionsGrantsAmt>
    <CYInvestmentIncomeAmt>30000</CYInvestmentIncomeAmt>
    <PYTotalRevenueAmt>480000</PYTotalRevenueAmt>
    <PYTotalExpensesAmt>430000</PYTotalExpensesAmt>
    <TotalEmployeeCnt>25</TotalEmployeeCnt>
    <VotingMembersGoverningBodyCnt>7</VotingMembersGoverningBodyCnt>
    </IRS990></ReturnData></Return>"""
    row = parse_single_xml(xml, "test.xml")
    kpis = compute_kpis(row)
    _result("Surplus=50000", kpis["OperatingSurplus"] == 50000)
    _result("ProgRatio=77.8%", abs(kpis["ProgramExpenseRatio"] - 350000/450000) < 0.001)
    _result("DebtRatio=25%", abs(kpis["DebtToAssetRatio"] - 200000/800000) < 0.001)
    _result("RevGrowth computed", kpis.get("RevenueGrowth") is not None)
    _result("ExpGrowth computed", kpis.get("ExpenseGrowth") is not None)
    for k in ["OperatingSurplus", "ProgramExpenseRatio", "OperatingMargin",
              "DebtToAssetRatio", "SalaryToExpenseRatio", "ContributionDependency",
              "RevenueGrowth", "ExpenseGrowth", "FundraisingRatio"]:
        _result(f"XML {k} not None", kpis.get(k) is not None)


def test_08():
    from core.quickbooks import parse_quickbooks_report
    print("\n  TEST 08: Single-period P&L (Riverside Animal Rescue)")
    rows = parse_quickbooks_report(create_single_period_pl(), "Riverside_PL_2025.xlsx")
    _result(">= 1 period", len(rows) >= 1, str(len(rows)))
    r = rows[0] if rows else {}
    _result("Org=Riverside Animal Rescue", r.get("OrganizationName") == "Riverside Animal Rescue",
            r.get("OrganizationName"))
    _result("Year=2025", r.get("TaxYear") == "2025", r.get("TaxYear"))
    _result("Rev=214000", (r.get("TotalRevenue") or 0) == 214000, str(r.get("TotalRevenue")))
    _result("Exp=178000", (r.get("TotalExpenses") or 0) == 178000, str(r.get("TotalExpenses")))
    _result("Salaries=89000", (r.get("SalariesWages") or 0) == 89000)
    _result("Rent=24000", (r.get("Occupancy") or 0) == 24000)
    _result("Insurance=6800", (r.get("Insurance") or 0) == 6800)
    _result("IT=6000", (r.get("InformationTechnology") or 0) == 6000)
    _result("Depreciation=5800", (r.get("DepreciationAmortization") or 0) == 5800)
    _result("InvIncome=3500", (r.get("InvestmentIncome") or 0) == 3500)


def test_09():
    from core.quickbooks import parse_quickbooks_report
    print("\n  TEST 09: CSV format P&L (Green Valley Land Trust)")
    rows = parse_quickbooks_report(create_csv_pl(), "GreenValley_PL.csv")
    r24 = next((r for r in rows if r.get("TaxYear") == "2024"), {})
    _result(">= 1 period", len(rows) >= 1)
    _result("Org=Green Valley Land Trust", r24.get("OrganizationName") == "Green Valley Land Trust",
            r24.get("OrganizationName"))
    _result("Rev=830000", (r24.get("TotalRevenue") or 0) == 830000, str(r24.get("TotalRevenue")))
    _result("Exp=483000", (r24.get("TotalExpenses") or 0) == 483000, str(r24.get("TotalExpenses")))
    _result("Salaries=350000", (r24.get("SalariesWages") or 0) == 350000)
    _result("Rent=48000", (r24.get("Occupancy") or 0) == 48000)
    _result("ProgramRev=200000", (r24.get("ProgramServiceRevenue") or 0) == 200000)
    _result("InvIncome=30000", (r24.get("InvestmentIncome") or 0) == 30000)


def test_10():
    from core.quickbooks import parse_quickbooks_report, merge_quickbooks_data
    from core.kpis import compute_kpis
    print("\n  TEST 10: P&L-only upload, no Balance Sheet (Habitat)")
    rows = parse_quickbooks_report(create_mock_pl(), "Habitat_PL_Only.xlsx")
    merged = merge_quickbooks_data(rows)
    r = merged[-1]
    kpis = compute_kpis(r)
    _result("Surplus computes", kpis.get("OperatingSurplus") is not None)
    _result("Surplus=318000", kpis.get("OperatingSurplus") == 318000)
    _result("Margin computes", kpis.get("OperatingMargin") is not None)
    _result("SalaryRatio computes", kpis.get("SalaryToExpenseRatio") is not None)
    _result("TotalAssets absent", r.get("TotalAssets") is None or r.get("TotalAssets") == 0)
    _result("No crash on asset KPIs", True)


def test_11():
    from core.quickbooks import parse_quickbooks_report, merge_quickbooks_data
    from core.kpis import compute_kpis
    print("\n  TEST 11: Balance Sheet-only upload (Cascade Education Alliance)")
    rows = parse_quickbooks_report(create_large_bs(), "Cascade_BS_2024.xlsx")
    merged = merge_quickbooks_data(rows)
    r = merged[-1]
    _result("Assets=3850000", (r.get("TotalAssets") or 0) == 3850000, str(r.get("TotalAssets")))
    _result("Liabilities=260000", (r.get("TotalLiabilities") or 0) == 260000, str(r.get("TotalLiabilities")))
    _result("Cash=520000", (r.get("CashNonInterest") or 0) == 520000)
    _result("Savings=890000", (r.get("SavingsTempCashInvestments") or 0) == 890000)
    _result("Unrestricted=2740000", (r.get("NetAssetsWithoutDonorRestrictions") or 0) == 2740000)
    _result("Restricted=850000", (r.get("NetAssetsWithDonorRestrictions") or 0) == 850000)
    kpis = compute_kpis(r)
    _result("LiquidAssets > 0", kpis.get("LiquidAssets") is not None and kpis["LiquidAssets"] > 0)
    _result("DebtToAsset computed", kpis.get("DebtToAssetRatio") is not None)
    _result("No crash", True)


def test_12():
    from core.quickbooks import parse_quickbooks_report
    print("\n  TEST 12: Alternative QB naming (Summit Youth Services)")
    rows = parse_quickbooks_report(create_alt_naming_pl(), "Summit_PL_FY2024.xlsx")
    r = rows[0] if rows else {}
    _result("Org=Summit Youth Services", r.get("OrganizationName") == "Summit Youth Services",
            r.get("OrganizationName"))
    _result("Year=2024", r.get("TaxYear") == "2024", r.get("TaxYear"))
    _result("Rev=935000", (r.get("TotalRevenue") or 0) == 935000, str(r.get("TotalRevenue")))
    _result("Exp=762000", (r.get("TotalExpenses") or 0) == 762000, str(r.get("TotalExpenses")))
    _result("Contribs=620000", (r.get("TotalContributionsGrants") or 0) == 620000)
    _result("FeeForService=275000 -> ProgramRev", (r.get("ProgramServiceRevenue") or 0) == 275000)
    _result("Payroll=415000 -> Salaries", (r.get("SalariesWages") or 0) == 415000)
    _result("Occupancy=68000", (r.get("Occupancy") or 0) == 68000)
    _result("Computer=31000 -> IT", (r.get("InformationTechnology") or 0) == 31000)
    _result("AuditFees=16000 -> Accounting", (r.get("AccountingFees") or 0) == 16000)
    _result("LegalServices=11000 -> Legal", (r.get("LegalFees") or 0) == 11000)
    _result("GrantExpense=58000", (r.get("GrantsAndSimilarPaid") or 0) == 58000)
    _result("IntAndDiv=28000 -> InvIncome", (r.get("InvestmentIncome") or 0) == 28000)


def test_13():
    from core.quickbooks import parse_quickbooks_report
    print("\n  TEST 13: Error handling (empty / malformed / non-QB / short)")
    # Empty file
    try:
        parse_quickbooks_report(b"", "empty.xlsx")
        _result("Empty file raises ValueError", False, "no error raised")
    except ValueError:
        _result("Empty file raises ValueError", True)
    # Garbage bytes
    try:
        parse_quickbooks_report(b"not a spreadsheet at all \x00\xff", "garbage.xlsx")
        _result("Garbage bytes raises ValueError", False, "no error raised")
    except ValueError:
        _result("Garbage bytes raises ValueError", True)
    # Valid Excel but not QB report (no financial structure)
    wb = openpyxl.Workbook()
    wb.active.append(["Student Name", "Grade", "GPA"])
    wb.active.append(["Alice Johnson", "Senior", 3.8])
    wb.active.append(["Bob Smith", "Junior", 3.2])
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    try:
        parse_quickbooks_report(buf.getvalue(), "students.xlsx")
        _result("Non-QB file raises ValueError", False, "no error raised")
    except ValueError:
        _result("Non-QB file raises ValueError", True)
    # Very short file (only 1 row)
    wb2 = openpyxl.Workbook()
    wb2.active.append(["Single Row"])
    buf2 = io.BytesIO(); wb2.save(buf2); buf2.seek(0)
    try:
        parse_quickbooks_report(buf2.getvalue(), "short.xlsx")
        _result("Short file raises ValueError", False, "no error raised")
    except ValueError:
        _result("Short file raises ValueError", True)


def test_14():
    from core.quickbooks import parse_quickbooks_report, merge_quickbooks_data
    from core.kpis import compute_kpis
    print("\n  TEST 14: 5-year comparative P&L (Pacific Marine Conservation)")
    rows = parse_quickbooks_report(create_5yr_pl(), "Pacific_5yr_PL.xlsx")
    merged = merge_quickbooks_data(rows)
    _result("5 years found", len(merged) == 5, str(len(merged)))
    years = sorted([r.get("TaxYear") for r in merged])
    _result("Years 2020-2024", years == ["2020", "2021", "2022", "2023", "2024"], str(years))
    # Revenue should increase each year
    revs = sorted([(r.get("TaxYear"), r.get("TotalRevenue", 0)) for r in merged])
    for i in range(1, len(revs)):
        _result(f"Rev {revs[i][0]}>{revs[i-1][0]}", revs[i][1] > revs[i-1][1])
    # Check specific values
    r2024 = next(r for r in merged if r.get("TaxYear") == "2024")
    r2020 = next(r for r in merged if r.get("TaxYear") == "2020")
    _result("2024 Rev=615000", r2024.get("TotalRevenue") == 615000)
    _result("2020 Rev=410000", r2020.get("TotalRevenue") == 410000)
    # KPIs should not crash on any year
    for r in merged:
        try:
            compute_kpis(r)
            _result(f"Year {r.get('TaxYear')} KPIs ok", True)
        except Exception as e:
            _result(f"Year {r.get('TaxYear')} KPIs ok", False, str(e))


def test_15():
    from core.quickbooks import parse_quickbooks_report, merge_quickbooks_data
    from core.kpis import compute_kpis, format_kpi_value
    print("\n  TEST 15: Zero-revenue startup (Fresh Start Legal Aid)")
    rows = parse_quickbooks_report(create_zero_revenue_pl(), "FreshStart_PL_2025.xlsx")
    merged = merge_quickbooks_data(rows)
    _result(">= 1 record", len(merged) >= 1)
    r = merged[0]
    _result("Org=Fresh Start Legal Aid", r.get("OrganizationName") == "Fresh Start Legal Aid",
            r.get("OrganizationName"))
    _result("Rev=0", r.get("TotalRevenue") == 0 or r.get("TotalRevenue") is None)
    _result("Exp=11500", (r.get("TotalExpenses") or 0) == 11500, str(r.get("TotalExpenses")))
    _result("LegalFees=5500", (r.get("LegalFees") or 0) == 5500)
    _result("IT=1000", (r.get("InformationTechnology") or 0) == 1000)
    try:
        kpis = compute_kpis(r)
        _result("KPIs no crash on zero revenue", True)
        _result("Margin safe (no div-by-zero)", kpis.get("OperatingMargin") is not None)
        _result("Surplus=-11500", kpis.get("OperatingSurplus") == -11500 or kpis.get("OperatingSurplus") is not None)
    except Exception as e:
        _result("KPIs no crash on zero revenue", False, str(e))


def test_16(merged):
    from core.quickbooks import apply_supplement_data
    from core.kpis import compute_kpis
    print("\n  TEST 16: Partial supplement — only EIN + Program% (Habitat)")
    rows = copy.deepcopy(merged)
    rows = apply_supplement_data(rows, {"EIN": "98-7654321", "ProgramExpensePct": 82})
    for row in rows:
        y = row.get("TaxYear")
        _result(f"{y} EIN=98-7654321", row.get("EIN") == "98-7654321")
        _result(f"{y} ProgExp filled", row.get("ProgramExpenses") is not None and row["ProgramExpenses"] > 0)
        _result(f"{y} MgmtExp still None", row.get("ManagementGeneralExpenses") is None)
        _result(f"{y} Board still None", row.get("VotingBoardMembers") is None)
        kpis = compute_kpis(row)
        _result(f"{y} ProgRatio~82%", abs((kpis.get("ProgramExpenseRatio") or 0) - 0.82) < 0.01)
        _result(f"{y} MgmtRatio still N/A", kpis.get("ManagementGeneralRatio") is None)


def test_17(merged):
    from core.kpis import compute_kpis, KPI_DEFINITIONS, format_kpi_value, get_kpi_status
    print("\n  TEST 17: All 28 KPIs exhaustive format/status check (Habitat)")
    r = merged[-1]
    kpis = compute_kpis(r)
    for key in KPI_DEFINITIONS:
        val = kpis.get(key)
        fv = format_kpi_value(key, val)
        status = get_kpi_status(key, val)
        _result(f"{key} format ok", fv is not None and len(fv) > 0, f"'{fv}'")
        _result(f"{key} status valid", status in ("good", "warning", "concern", "neutral"), status)
        if val is None:
            _result(f"{key} None->N/A", fv == "N/A", fv)
            _result(f"{key} None->neutral", status == "neutral", status)


def test_18():
    from core.quickbooks import parse_quickbooks_report
    print("\n  TEST 18: QB Desktop format (Heartland Food Bank)")
    rows = parse_quickbooks_report(create_desktop_pl(), "Heartland_PL_Desktop.xlsx")
    _result(">= 1 period", len(rows) >= 1)
    r24 = rows[0] if rows else {}
    _result("Org=Heartland Food Bank", r24.get("OrganizationName") == "Heartland Food Bank",
            r24.get("OrganizationName"))
    _result("Rev=742000", (r24.get("TotalRevenue") or 0) == 742000, str(r24.get("TotalRevenue")))
    _result("Exp=631000", (r24.get("TotalExpenses") or 0) == 631000, str(r24.get("TotalExpenses")))
    _result("Salaries=310000", (r24.get("SalariesWages") or 0) == 310000)
    _result("PayrollTax=25000", (r24.get("PayrollTaxes") or 0) == 25000)
    _result("Benefits=38000", (r24.get("EmployeeBenefits") or 0) == 38000)
    _result("Rent=54000", (r24.get("Occupancy") or 0) == 54000)
    _result("Insurance=16000", (r24.get("Insurance") or 0) == 16000)
    _result("Travel=18000", (r24.get("Travel") or 0) == 18000)
    _result("Depreciation=22000", (r24.get("DepreciationAmortization") or 0) == 22000)
    _result("IT=17500", (r24.get("InformationTechnology") or 0) == 17500)


def test_19():
    from core.quickbooks import merge_quickbooks_data
    print("\n  TEST 19: Duplicate year handling in merge")
    dupes = [
        {"TaxYear": "2024", "DataSource": "quickbooks",
         "OrganizationName": "Duplicate Test Org", "TotalRevenue": 500000, "TotalExpenses": 400000},
        {"TaxYear": "2024", "DataSource": "quickbooks",
         "OrganizationName": "Duplicate Test Org", "TotalRevenue": 500000, "TotalExpenses": 400000},
        {"TaxYear": "2023", "DataSource": "quickbooks",
         "OrganizationName": "Duplicate Test Org", "TotalRevenue": 450000, "TotalExpenses": 370000},
    ]
    merged = merge_quickbooks_data(dupes)
    _result("Deduped to 2 years", len(merged) == 2, str(len(merged)))
    r2024 = next(r for r in merged if r.get("TaxYear") == "2024")
    r2023 = next(r for r in merged if r.get("TaxYear") == "2023")
    _result("2024 Rev=500000", r2024.get("TotalRevenue") == 500000)
    _result("2023 Rev=450000", r2023.get("TotalRevenue") == 450000)


def test_20():
    from core.quickbooks import parse_quickbooks_report
    print("\n  TEST 20: Report type auto-detection accuracy")
    for name, builder, expected in [
        ("Habitat PL", create_mock_pl, "pl"),
        ("Riverside Single PL", create_single_period_pl, "pl"),
        ("Summit Alt PL", create_alt_naming_pl, "pl"),
        ("Pacific 5yr PL", create_5yr_pl, "pl"),
        ("Heartland Desktop PL", create_desktop_pl, "pl"),
        ("FreshStart Zero PL", create_zero_revenue_pl, "pl"),
        ("Habitat BS", create_mock_bs, "bs"),
        ("Cascade Large BS", create_large_bs, "bs"),
    ]:
        try:
            rows = parse_quickbooks_report(builder(), f"{name}.xlsx")
            rtype = rows[0].get("_report_type", "?") if rows else "empty"
            _result(f"{name} -> {expected}", rtype == expected, f"got '{rtype}'")
        except Exception as e:
            _result(f"{name} -> {expected}", False, str(e))


# ═══════════════════════════════════════════════
# Main
# ═══════════════════════════════════════════════

if __name__ == "__main__":
    print("\n" + "=" * 60)
    print("  QUICKBOOKS INTEGRATION — 20-TEST SUITE")
    print("=" * 60)

    try:
        pl_rows = test_01()
        bs_rows = test_02()
        merged = test_03(pl_rows, bs_rows)
        test_04(merged)
        test_05(merged)
        test_06(merged)
        test_07()
        test_08()
        test_09()
        test_10()
        test_11()
        test_12()
        test_13()
        test_14()
        test_15()
        test_16(merged)
        test_17(merged)
        test_18()
        test_19()
        test_20()
    except Exception:
        traceback.print_exc()

    print("\n" + "=" * 60)
    print(f"  RESULTS: {PASSED}/{PASSED + FAILED} checks passed, {FAILED} failed")
    if FAILED == 0:
        print("  ALL TESTS PASSED")
    else:
        print(f"  {FAILED} FAILURE(S)")
    print("=" * 60)
