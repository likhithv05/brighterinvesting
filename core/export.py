"""
Excel Export Module
Generates professionally formatted Excel workbooks from parsed Form 990 data.

Developed by Epic Intentions for Brighter Investing
Georgia Institute of Technology — Spring 2026
"""

import io
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from core.kpis import compute_kpis, KPI_DEFINITIONS


# Color palette - Brighter Investing Brand (Light theme)
COLORS = {
    "header_bg": "0F172A",       # Navy (professional)
    "header_font": "FFFFFF",     # White text on dark header
    "subheader_bg": "F8FAFC",   # Light slate background
    "subheader_font": "0F172A",  # Dark text
    "accent": "0D9488",          # Teal (brand primary)
    "good": "059669",            # Status green
    "warning": "D97706",         # Status amber
    "concern": "E11D48",         # Status red
    "light_row": "F8FAFC",      # Very subtle gray tint
    "border": "E2E8F0",          # Light border
}


def _style_header_row(ws, max_col, row=1):
    """Apply professional header styling to a row."""
    header_font = Font(name="Calibri", bold=True, color=COLORS["header_font"], size=11)
    header_fill = PatternFill(start_color=COLORS["header_bg"], end_color=COLORS["header_bg"], fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        bottom=Side(style="thin", color=COLORS["border"]),
        right=Side(style="thin", color=COLORS["border"]),
    )
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border


def _auto_width(ws, min_width=12, max_width=30):
    """Auto-size column widths based on content."""
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_len = min_width
        for row_idx in range(1, min(ws.max_row + 1, 50)):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value:
                max_len = max(max_len, min(len(str(cell.value)) + 2, max_width))
        ws.column_dimensions[col_letter].width = max_len


def _apply_zebra_striping(ws, start_row=2):
    """Apply alternating row colors for readability."""
    light_fill = PatternFill(start_color=COLORS["light_row"], end_color=COLORS["light_row"], fill_type="solid")
    for row_idx in range(start_row, ws.max_row + 1):
        if row_idx % 2 == 0:
            for col_idx in range(1, ws.max_column + 1):
                ws.cell(row=row_idx, column=col_idx).fill = light_fill


def _format_currency_cells(ws, columns, start_row=2):
    """Apply currency number format to specified columns."""
    for row_idx in range(start_row, ws.max_row + 1):
        for col_idx in columns:
            cell = ws.cell(row=row_idx, column=col_idx)
            if isinstance(cell.value, (int, float)):
                cell.number_format = '$#,##0;[Red]-$#,##0'


def _format_percent_cells(ws, columns, start_row=2):
    """Apply percentage number format to specified columns."""
    for row_idx in range(start_row, ws.max_row + 1):
        for col_idx in columns:
            cell = ws.cell(row=row_idx, column=col_idx)
            if isinstance(cell.value, (int, float)):
                cell.number_format = '0.0%'


def generate_workbook(parsed_rows):
    """
    Generate a professionally formatted Excel workbook from parsed Form 990 data.

    Parameters:
        parsed_rows: list of dicts from parse_single_xml()

    Returns:
        bytes (Excel file content)
    """
    wb = openpyxl.Workbook()

    # =========================================================
    # SHEET 1: Summary Dashboard
    # =========================================================
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary.sheet_properties.tabColor = COLORS["accent"]

    # Title row
    ws_summary.merge_cells("A1:H1")
    title_cell = ws_summary.cell(row=1, column=1, value="Form 990 Financial Analysis")
    title_cell.font = Font(name="Calibri", bold=True, size=16, color=COLORS["header_bg"])
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws_summary.row_dimensions[1].height = 35

    # Subtitle
    ws_summary.merge_cells("A2:H2")
    if parsed_rows:
        org_name = parsed_rows[0].get("OrganizationName", "Unknown Organization")
        years = sorted([r.get("TaxYear", "") for r in parsed_rows])
        year_range = f"{years[0]} – {years[-1]}" if len(years) > 1 else years[0]
        subtitle = f"{org_name}  |  Tax Years: {year_range}"
    else:
        subtitle = "No data loaded"
    sub_cell = ws_summary.cell(row=2, column=1, value=subtitle)
    sub_cell.font = Font(name="Calibri", size=11, color="6B7280", italic=True)
    ws_summary.row_dimensions[2].height = 22

    # Attribution
    ws_summary.merge_cells("A3:H3")
    attr_cell = ws_summary.cell(row=3, column=1, value="Prepared by Epic Intentions | Brighter Investing | Georgia Tech")
    attr_cell.font = Font(name="Calibri", size=9, color="9CA3AF")
    ws_summary.row_dimensions[3].height = 18

    # KPI Table
    kpi_start_row = 5
    kpi_headers = ["Tax Year"]
    kpi_keys_ordered = [
        "OperatingSurplus", "TotalCashEquivalents", "ProgramExpenseRatio",
        "ManagementGeneralRatio", "FundraisingRatio", "OperatingMargin",
        "ContributionDependency", "LiquidAssets", "MonthsExpenseCoverage",
        "NetOperatingIncome", "NetInvestmentGain", "DepreciationNonCash",
        "DebtToAssetRatio", "CurrentRatio", "RevenueGrowth",
        "ExpenseGrowth", "SalaryToExpenseRatio", "NetAssetGrowth",
    ]

    for key in kpi_keys_ordered:
        defn = KPI_DEFINITIONS.get(key, {})
        kpi_headers.append(defn.get("label", key))

    # Write headers
    for i, h in enumerate(kpi_headers, start=1):
        ws_summary.cell(row=kpi_start_row, column=i, value=h)
    _style_header_row(ws_summary, len(kpi_headers), row=kpi_start_row)

    # Sort rows by tax year
    sorted_rows = sorted(parsed_rows, key=lambda r: r.get("TaxYear", ""))

    # Write KPI data rows
    currency_cols = []
    percent_cols = []
    for row_idx, row in enumerate(sorted_rows, start=kpi_start_row + 1):
        kpis = compute_kpis(row)
        ws_summary.cell(row=row_idx, column=1, value=row.get("TaxYear", ""))
        ws_summary.cell(row=row_idx, column=1).font = Font(name="Calibri", bold=True, size=11)

        for col_offset, key in enumerate(kpi_keys_ordered, start=2):
            raw_val = kpis.get(key, 0)
            display_val = "N/A" if raw_val is None else raw_val
            cell = ws_summary.cell(row=row_idx, column=col_offset, value=display_val)
            cell.font = Font(name="Calibri", size=10)
            cell.alignment = Alignment(horizontal="right")

            fmt = KPI_DEFINITIONS.get(key, {}).get("format", "currency")
            if fmt == "currency":
                cell.number_format = '$#,##0;[Red]-$#,##0'
                if col_offset not in currency_cols:
                    currency_cols.append(col_offset)
            elif fmt == "percent":
                cell.number_format = '0.0%'
                if col_offset not in percent_cols:
                    percent_cols.append(col_offset)
            elif fmt == "decimal":
                cell.number_format = '0.0'
            elif fmt == "ratio":
                cell.number_format = '0.00"x"'

    _apply_zebra_striping(ws_summary, start_row=kpi_start_row + 1)
    _auto_width(ws_summary, min_width=14, max_width=28)
    ws_summary.freeze_panes = f"B{kpi_start_row + 1}"

    # =========================================================
    # SHEET 2: Raw Extracted Data (Inputs)
    # =========================================================
    ws_inputs = wb.create_sheet("Extracted Data")
    ws_inputs.sheet_properties.tabColor = "0891B2"

    input_fields = [
        "SourceFile", "OrganizationName", "EIN", "TaxYear",
        "TotalRevenue", "PriorYearRevenue", "GrossReceipts",
        "TotalContributionsGrants", "GovernmentGrants",
        "ContributionsGrantsOther", "ProgramServiceRevenue",
        "InvestmentIncome", "OtherRevenue", "UnrelatedBusinessRevenue",
        "NetGainLossInvestments",
        "TotalExpenses", "PriorYearExpenses", "ProgramExpenses",
        "ManagementGeneralExpenses", "FundraisingExpenses",
        "SalariesWages", "PensionRetirementContributions",
        "EmployeeBenefits", "PayrollTaxes", "LegalFees",
        "AccountingFees", "OfficeExpenses", "InformationTechnology",
        "Occupancy", "Travel", "DepreciationAmortization",
        "Insurance", "GrantsAndSimilarPaid", "OtherExpenses",
        "CashNonInterest", "SavingsTempCashInvestments",
        "PublicInvestments", "Inventory", "PrepaidExpenses",
        "PropertyEquipmentNet", "OtherAssets", "TotalAssets",
        "AccountsPayableAccrued", "OtherLiabilities", "TotalLiabilities",
        "NetAssetsWithoutDonorRestrictions", "NetAssetsWithDonorRestrictions",
        "TotalNetAssets",
        "EmployeeCount", "Volunteers", "VotingBoardMembers",
        "IndependentBoardMembers", "ExecutiveDirectorCompensation",
        "DonatedServicesFacilities",
    ]

    from core.parser import FIELD_LABELS

    # Write headers with human-readable labels
    for i, field in enumerate(input_fields, start=1):
        label = FIELD_LABELS.get(field, field)
        ws_inputs.cell(row=1, column=i, value=label)
    _style_header_row(ws_inputs, len(input_fields))

    # Write data
    for row_idx, row in enumerate(sorted_rows, start=2):
        for col_idx, field in enumerate(input_fields, start=1):
            val = row.get(field, "")
            display_val = "N/A" if val is None else val
            ws_inputs.cell(row=row_idx, column=col_idx, value=display_val)
            ws_inputs.cell(row=row_idx, column=col_idx).font = Font(name="Calibri", size=10)

    # Apply currency formatting to numeric columns (columns 5 onward, except counts)
    count_fields = {"EmployeeCount", "Volunteers", "VotingBoardMembers", "IndependentBoardMembers"}
    currency_input_cols = []
    for i, field in enumerate(input_fields, start=1):
        if i >= 5 and field not in count_fields:
            currency_input_cols.append(i)
    _format_currency_cells(ws_inputs, currency_input_cols)
    _apply_zebra_striping(ws_inputs)
    _auto_width(ws_inputs, min_width=14, max_width=26)
    ws_inputs.freeze_panes = "E2"

    # =========================================================
    # SHEET 3: Benchmarks & Definitions
    # =========================================================
    ws_bench = wb.create_sheet("KPI Definitions")
    ws_bench.sheet_properties.tabColor = "C8E636"

    bench_headers = ["KPI", "Description", "Benchmark / Guidance"]
    for i, h in enumerate(bench_headers, start=1):
        ws_bench.cell(row=1, column=i, value=h)
    _style_header_row(ws_bench, 3)

    for row_idx, (key, defn) in enumerate(KPI_DEFINITIONS.items(), start=2):
        ws_bench.cell(row=row_idx, column=1, value=defn["label"])
        ws_bench.cell(row=row_idx, column=1).font = Font(name="Calibri", bold=True, size=10)
        ws_bench.cell(row=row_idx, column=2, value=defn["description"])
        ws_bench.cell(row=row_idx, column=2).font = Font(name="Calibri", size=10)
        ws_bench.cell(row=row_idx, column=2).alignment = Alignment(wrap_text=True)
        ws_bench.cell(row=row_idx, column=3, value=defn["benchmark"])
        ws_bench.cell(row=row_idx, column=3).font = Font(name="Calibri", size=10)
        ws_bench.cell(row=row_idx, column=3).alignment = Alignment(wrap_text=True)

    _apply_zebra_striping(ws_bench)
    ws_bench.column_dimensions["A"].width = 28
    ws_bench.column_dimensions["B"].width = 55
    ws_bench.column_dimensions["C"].width = 50

    # Save to bytes
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
